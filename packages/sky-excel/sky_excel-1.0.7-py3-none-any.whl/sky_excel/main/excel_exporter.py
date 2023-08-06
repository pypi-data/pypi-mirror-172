# encoding: utf-8
"""
@project: sky-excel->excel_exporter
@author: 孙楷炎
@Email: sky4834@163.com
@synopsis: excel输出者
@created_time: 2022/9/29 17:01
"""
from io import BytesIO
import os
import platform
import re
import uuid

from openpyxl import load_workbook

from .body_parser import BodyParser


class ExcelExport:
    input_dict = None
    header_cells = None
    body_map = None
    __err_msg = None
    wb = None

    @property
    def err_msg(self):
        return self.__err_msg

    @err_msg.setter
    def err_msg(self, err):
        # 返回最先暴出来的错误，按照顺序返回，避免覆盖错误无法排查
        if self.__err_msg is None:
            self.__err_msg = err

    def is_valid(self):
        return self.__err_msg

    def __init__(self, input_dict, excel_templet_path=None, excel_templet_title="Sheet1", save_path=None, ):
        # 直接解析模板，获取可操作的excel对象
        self.wb = load_workbook(filename=excel_templet_path) if excel_templet_path and os.path.exists(excel_templet_path) else None
        # body解析
        self.parsed_body_data, err = BodyParser(work_book_obj=self.wb).parse_body(sheet_name=excel_templet_title)
        if err:
            self.err_msg = err

        self.sheetnames = self.wb.sheetnames if self.wb else []  # 共工作簿列表
        self.sheetname = excel_templet_title
        self.save_path = save_path  # 当save_path 为空的时候则进行返回文件流的格式
        self.input_dict = input_dict
        # 合法性判断
        if not isinstance(input_dict, list):
            self.err_msg = "input_dict格式应该是[{..},{..}..]"
        if not excel_templet_path or not os.path.exists(excel_templet_path):
            self.err_msg = "excel 模板路径有误"
        if not self.sheetnames or not excel_templet_title in self.sheetnames:
            self.err_msg = "不存在名称为：%s 的工作簿" % excel_templet_title

    def write_body(self, sheet):
        try:
            # 可单独调用，不需要表头
            start_row = self.parsed_body_data.max_row or 1
            cell_list = self.parsed_body_data.cell_list or []  # 单元个对象
            # 写入数据 到数据体里面
            for row_num, row_input_dict in zip(range(start_row, len(self.input_dict) + start_row), self.input_dict):
                for cell in cell_list:
                    # 数据处理
                    this_cell_start_coord = re.compile("(\d+)").sub(str(row_num), cell.start_cell_coord)
                    this_cell_coordinate = re.compile("(\d+)").sub(str(row_num), cell.coordinate)
                    this_cell_value = cell.value

                    # 替换变量表达式
                    this_cell_value_match = re.compile("{{.*?}}").findall(str(cell.value) if cell.value else "")
                    this_cell_value_replaces = {i: row_input_dict.get(i.replace("{{", "").replace("}}", ""), "") for i in this_cell_value_match}
                    for k, v in this_cell_value_replaces.items():
                        this_cell_value = this_cell_value.replace(k, v)

                    # 数据写入
                    sheet[this_cell_start_coord] = this_cell_value
                    if cell.is_merge_cell:
                        sheet.merge_cells(this_cell_coordinate)
            return sheet, None
        except Exception as e:
            err_detail = " file:" + str(e.__traceback__.tb_frame.f_globals["__file__"]) + " line:" + str(e.__traceback__.tb_lineno) if platform.python_version()[0] == "3" else ""
            return sheet, "写入body异常：" + str(e) + err_detail

    # 整个流程的调度函数
    def export(self):
        if self.err_msg:
            return None, self.err_msg
        try:
            # 写入数据
            writed_body_sheet, body_err = self.write_body(self.wb[self.sheetname])
            # 数据写入异常
            if body_err:
                self.err_msg = body_err
                return None, body_err
            # 保存excel
            if self.save_path:
                try:
                    # 递归创建目录
                    if not os.path.exists(self.save_path):
                        os.makedirs(self.save_path)
                    # 保存文件
                    file_path_name = (self.save_path + "/" + str(uuid.uuid3(uuid.NAMESPACE_DNS, str(uuid.uuid1()))) + ".xlsx").replace("//", "/")
                    self.wb.save(file_path_name)
                    return file_path_name, None
                except Exception as e:
                    return None, "excel 写入异常：" + str(e)
            else:  # 没有路径则返回文件流的形式
                return self.export_stream(self.wb), None

        except PermissionError as pe:
            return None, "文件被占用，请关闭文件后在执行程序"
        except Exception as e:
            return None, "系统异常:" + str(e)

    # 返回文件流
    def export_stream(self, new_wb):
        try:
            from StringIO import StringIO
        except ImportError:
            from io import StringIO

        if platform.python_version()[0] == "2":
            # python 2.*.*
            output = StringIO()
            new_wb.save(output)
            output.seek(0)
            return output.getvalue()
        else:
            # python 3.*.*
            output = BytesIO()
            new_wb.save(output)
            output.seek(0)
            return output.getvalue()
