# encoding: utf-8
"""
@project: sky-excel->reader
@author: 孙楷炎
@Email: sky4834@163.com
@synopsis: excel模板解析器
@created_time: 2022/9/28 17:46
"""
import os
import re

from openpyxl import load_workbook
from openpyxl.cell import MergedCell

# 头部解析器
from .parsed_cell import ParsedCell


class HeaderParser(object):
    wb = None
    sheetnames = []
    __header = None

    @property
    def header(self):
        return self.__header

    @header.setter
    def header(self, cells):
        # TODO 把所有的单元格转换成自定义的类，
        #  因为合并单元表格与普通单元表格属性不一样，
        #  直接给出去出现异常判断问题，所以用以接口类，并且实现{sheet:cells}
        self.__header = [ParsedCell(i) for i in cells]

    # 获取工作工作簿实例
    def get_wb(self):
        return self.wb

    def __init__(self, file_path=None, work_book_obj=None):
        # excel对象
        if work_book_obj:
            self.wb = work_book_obj
        elif file_path and os.path.exists(file_path):
            self.wb = load_workbook(filename=file_path)
        # 共工作簿列表
        self.sheetnames = self.wb.sheetnames

    def parse_header(self, sheet_name="Sheet1"):
        try:
            assert not self.wb is None, "excel 初始化失败"
            assert sheet_name and sheet_name in self.sheetnames, "不存在名称为：%s 的工作簿" % sheet_name
            sheet = self.wb[sheet_name]
            patt = "{{.*}}"
            self.cells = sheet.merged_cell_ranges or []
            self.cells = [i for i in self.cells if not re.search(patt, str(i.start_cell.value), re.I)]  # 排除表达式符号{{}}合并的单元格
            start_cell_map = {i.start_cell.coordinate: i for i in self.cells}  # 合并单元格的开始单元格映射
            # 单元格处理
            for row in sheet.iter_rows():
                this_row_cells = []
                has_expression = False
                for cell in row:
                    # 是被合并的单元格，合并单元格的左上角的单元格的时候
                    if isinstance(cell, MergedCell) or start_cell_map.get(cell.coordinate):
                        continue
                    if re.search(patt, str(cell.value), re.I):
                        has_expression = True
                        continue
                    this_row_cells.append(cell)
                    assert not cell.value is None, "excel模板设置错误，不允许出现空值。单元格坐标：" + cell.coordinate
                # 当前行出现表达式的时候，停止解析
                if has_expression:
                    break
                self.cells += this_row_cells
            # 表头单元格处理
            self.header = self.cells
            return self.header, None
        except AssertionError as logic_err:
            return None, "逻辑错误：" + str(logic_err)
        except Exception as e:
            return None, "系统异常：" + str(e) + " file:" + str(e.__traceback__.tb_frame.f_globals["__file__"]) + " line:" + str(e.__traceback__.tb_lineno)
