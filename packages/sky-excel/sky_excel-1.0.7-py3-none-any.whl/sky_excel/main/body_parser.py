# encoding: utf-8
"""
@project: sky-excel->body_parser
@author: 孙楷炎
@Email: sky4834@163.com
@synopsis: 内容键值解析器
@created_time: 2022/9/29 14:24
"""

import os
import platform

from openpyxl import load_workbook
# 头部解析器
from openpyxl.cell import MergedCell

from .parsed_cell import ParsedCell
from ..utils.j_dict import JDict


class BodyParser(object):
    wb = None
    __cell_list = []

    @property
    def cell_list(self):
        return self.__cell_list

    @cell_list.setter
    def cell_list(self, cells):
        self.__cell_list = [ParsedCell(i, self.sheet) for i in cells]

    # 获取工作工作簿实例
    def get_wb(self):
        return self.wb

    def __init__(self, file_path=None, work_book_obj=None):
        # excel对象
        if work_book_obj:
            self.wb = work_book_obj
        elif file_path and os.path.exists(file_path):
            self.wb = load_workbook(filename=file_path)
        # 共工作簿列表
        self.sheetnames = self.wb.sheetnames if self.wb else None

    def parse_body(self, sheet_name="Sheet1"):
        try:
            assert not self.wb is None, "excel 初始化失败，检出模板文件路径"
            assert sheet_name and sheet_name in self.sheetnames, "不存在名称为：%s 的工作簿" % sheet_name
            self.sheet = self.wb[sheet_name]

            # 获取合并的单元格其实单元格映射
            cells = [i for i in self.sheet.merged_cell_ranges or [] if i.max_row == self.sheet.max_row]
            # 合并单元格的开始单元格映射,版本兼容
            if platform.python_version()[0] == "3":
                start_cell_map = {i.start_cell.coordinate: i for i in cells}
            else:
                start_cell_map = {self.sheet[i.coord][0][0].coordinate: i for i in cells}

            # 最后一行代表公式行
            for cell in self.sheet[self.sheet.max_row]:
                if isinstance(cell, MergedCell) or start_cell_map.get(cell.coordinate):
                    continue
                cells.append(cell)
            # 单元格进一步封装成通用格式
            self.cell_list = cells
            return JDict({"cell_list": self.cell_list, "max_row": self.sheet.max_row}), None
        except AssertionError as logic_err:
            return None, "逻辑错误：" + str(logic_err)
        except Exception as e:
            err_detail = " file:" + str(e.__traceback__.tb_frame.f_globals["__file__"]) + " line:" + str(e.__traceback__.tb_lineno) if platform.python_version()[0] == "3" else ""
            return None, "系统异常：" + str(e) + err_detail
