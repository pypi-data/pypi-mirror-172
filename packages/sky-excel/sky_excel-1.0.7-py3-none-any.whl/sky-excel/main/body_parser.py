# encoding: utf-8
"""
@project: sky-excel->body_parser
@author: 孙楷炎
@Email: sky4834@163.com
@synopsis: 内容键值解析器
@created_time: 2022/9/29 14:24
"""

import os
import re

from openpyxl import load_workbook
from openpyxl.cell import MergedCell
# 头部解析器
from openpyxl.worksheet.merge import MergedCellRange


class BodyParser(object):
    wb = None
    sheetnames = []
    __body_map = {}

    @property
    def body_map(self):
        return self.__body_map

    @body_map.setter
    def body_map(self, cells):
        self.__body_map = {}
        for cell in cells:
            if isinstance(cell, MergedCellRange):
                temp_cell = {str(cell): str(cell.start_cell.value).replace("{{", "").replace("}}", "")}
            else:
                temp_cell = {str(cell.coordinate): str(cell.value).replace("{{", "").replace("}}", "")}
            self.__body_map.update(temp_cell)

        # self.__body_map = [ParsedCell(i) for i in cells]

    # 获取工作工作簿实例
    def get_wb(self):
        return self.wb

    def __init__(self, file_path=None, work_book_obj=None):
        # excel对象
        if work_book_obj:
            self.wb = work_book_obj
        elif file_path and os.path.exists(file_path):
            self.wb = load_workbook(filename=file_path)
        # 共工作簿列表
        self.sheetnames = self.wb.sheetnames

    def parse_body(self, sheet_name="Sheet1"):
        try:
            assert not self.wb is None, "excel 初始化失败，检出模板文件路径"
            assert sheet_name and sheet_name in self.sheetnames, "不存在名称为：%s 的工作簿" % sheet_name
            sheet = self.wb[sheet_name]
            patt = "{{.*}}"
            self.cells = sheet.merged_cell_ranges or []
            self.cells = [i for i in self.cells if re.search(patt, str(i.start_cell.value), re.I)]  # 表达式符号{{}}合并的单元格
            start_cell_map = {c.start_cell.coordinate: c for c in self.cells}  # 合并单元格的开始单元格映射
            # 单元格处理
            for row, row_num in zip(sheet.iter_rows(), range(sheet.max_row)):
                if not row_num + 1 == sheet.max_row:
                    continue
                for cell in row:
                    if isinstance(cell, MergedCell) or start_cell_map.get(cell.coordinate):
                        continue
                    assert not cell.value is None, "模板错误，body映射表不允许出现空值，应该是{{}}符号，里面可以是 “{{=expression}}”、“{{key}}”。位置：" + cell.coordinate
                    self.cells.append(cell)
            # 表头单元格处理
            self.body_map = self.cells
            return {"map": self.body_map, "max_row": sheet.max_row}, None
        except AssertionError as logic_err:
            return None, "逻辑错误：" + str(logic_err)
        except Exception as e:
            return None, "系统异常：" + str(e) + " file:" + str(e.__traceback__.tb_frame.f_globals["__file__"]) + " line:" + str(e.__traceback__.tb_lineno)
