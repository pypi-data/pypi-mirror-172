# encoding: utf-8
"""
@project: sky-excel->excel_exporter
@author: 孙楷炎
@Email: sky4834@163.com
@synopsis: excel输出者
@created_time: 2022/9/29 17:01
"""
from io import BytesIO
import os
import platform
import re
import uuid

from openpyxl import Workbook

from .body_parser import BodyParser
from .header_parser import HeaderParser


class ExcelExport:
    input_dict = None
    header_cells = None
    body_map = None
    __err_msg = None

    def is_valid(self):
        return self.__err_msg

    def __init__(self, input_dict, excel_templet_path=None, excel_templet_title="Sheet1", save_path=None, ):
        self.save_path = save_path  # 当save_path 为空的时候则进行返回文件流的格式
        self.input_dict = input_dict
        if not isinstance(input_dict, list):
            self.__err_msg = "input_dict格式应该是[{..},{..}..]"
        # 头部解析
        header_parser = HeaderParser(file_path=excel_templet_path)
        self.header_cells, err = header_parser.parse_header(sheet_name=excel_templet_title)
        if err:
            self.err_msg = err

        # body解析
        body_parser = BodyParser(work_book_obj=header_parser.get_wb())
        self.body_map, err = body_parser.parse_body(sheet_name=excel_templet_title)
        if err:
            self.__err_msg = err

    def write_header(self, sheet):
        try:
            # 抄写模板的头部
            for header_cell in self.header_cells:
                if header_cell.is_merge_cell:
                    sheet.merge_cells(header_cell.coordinate)
                    sheet.cell(row=header_cell.row, column=header_cell.column, value=header_cell.value)
                else:
                    sheet[header_cell.coordinate] = str(header_cell.value)
            return sheet, None
        except Exception as e:
            return sheet, "写入头部异常：" + str(e)

    def write_body(self, sheet):
        try:
            # 可单独调用，不需要表头
            start_rows = self.body_map.get("max_row")
            cell_to_key = self.body_map.get("map")  # {A1:key} ，坐标列对应键名
            key_to_cell = {v: k for k, v in cell_to_key.items()}  # key:A1, 数据键名对应坐标列
            # 写入数据 到数据体里面
            for row_data_dict, this_row_num in zip(self.input_dict, range(start_rows, len(self.input_dict) + start_rows)):
                for k, v in row_data_dict.items():
                    is_has_key = key_to_cell.get(k)  # 得到A1或者A1:H1,所以判断拆分
                    # 边界判断 如果没有找到该键名称跳过
                    if not is_has_key:
                        continue
                    cell_index = is_has_key.split(":")
                    if len(cell_index) > 1:  #
                        start_out = re.sub("\d*", "", cell_index[0]) + str(this_row_num)
                        end_out = re.sub("\d*", "", cell_index[1]) + str(this_row_num)
                        sheet[start_out].value = str(v)
                        sheet.merge_cells(start_out + ":" + end_out)
                    else:
                        single_out = re.sub("\d*", "", cell_index[0]) + str(this_row_num)
                        sheet[single_out].value = str(v)

            return sheet, None
        except Exception as e:
            return sheet, "写入body异常：" + str(e)

    # 返回文件流
    def export_stream(self, new_wb):
        try:
            from StringIO import StringIO
        except ImportError:
            from io import StringIO

        if platform.python_version()[0] == "2":
            # python 2.*.*
            output = StringIO()
            new_wb.save(output)
            output.seek(0)
            return output.getvalue()
        else:
            # python 3.*.*
            output = BytesIO()
            new_wb.save(output)
            output.seek(0)
            return output.getvalue()

    # 整个流程的调度函数
    def export(self, export_sheet_name="Sheet1"):
        if self.__err_msg:
            return None, self.__err_msg
        try:
            # 创建新的工作簿
            new_wb = Workbook()
            export_sheet = new_wb.active
            export_sheet.title = export_sheet_name

            # 写入数据
            writed_header_sheet, header_err = self.write_header(new_wb[export_sheet_name])
            writed_body_sheet, body_err = self.write_body(writed_header_sheet)
            # 数据写入异常
            if body_err:
                return None, body_err
            if header_err:
                return None, header_err

            # 保存excel
            if self.save_path:
                try:
                    # 递归创建目录
                    if not os.path.exists(self.save_path):
                        os.makedirs(self.save_path)
                    # 保存文件
                    file_path_name = (self.save_path + "/" + str(uuid.uuid3(uuid.NAMESPACE_DNS, str(uuid.uuid1()))) + ".xlsx").replace("//", "/")
                    new_wb.save(file_path_name)
                    return file_path_name, None
                except Exception as e:
                    return None, "excel 写入异常：" + str(e)
            else:
                return self.export_stream(new_wb), None
        except PermissionError:
            return None, "文件被占用，请关闭文件后在执行程序"
        except Exception as e:
            return None, "系统异常:" + str(e)
